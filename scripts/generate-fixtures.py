#!/usr/bin/env python3
"""生成脱敏测试模板 tests/fixtures/立案与强执查询表-脱敏模板.xlsx。

结构复刻真实模板（立案与强执查询表-模板.xlsx）：
- Sheet1 单工作表，12 列 A-L；
- 立案表头第 1 行，数据第 2-4 行；
- 强执表头第 9 行（立案末行 4 + 5），数据第 10-12 行；
- 表头加粗 11 号、填充 FF92D050、行高 27；数据行高 28；
- 列宽 A15/B14/C20.37/D15.5/F13.25/G24.13/H12.87/J39.63/K18/L10.75；
- 日期列（F/I/L）数字格式 mm-dd-yy；
- H3 含一条 DISPIMG 公式（导入跳过逻辑测试）。

所有数据均为模拟值，不得出现真实业务数据。
用法: python scripts/generate-fixtures.py
"""
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "立案与强执查询表-脱敏模板.xlsx"

HEADER_LII = ["原告", "被告", "账号", "密码", "立案状态", "立案成功时间", "案号",
              "成功图片", "驳回时间", "驳回原因", "驳回图片", "查询时间"]
HEADER_QZ = ["原告", "被告", "账号", "密码", "强执状态", "强执成功时间", "强执案号",
             "成功图片", "驳回时间", "驳回原因", "驳回图片", "查询时间"]

COL_WIDTHS = {"A": 15.0, "B": 14.0, "C": 20.37, "D": 15.5, "E": 12.0, "F": 13.25,
              "G": 24.13, "H": 12.87, "I": 12.0, "J": 39.63, "K": 18.0, "L": 10.75}

# 模拟数据（显式伪造，禁止真实业务数据）
LII_ROWS = [
    ["测试原告甲", "测试被告A", "TEST-ACCOUNT-001", "test-pass-001", "审核中", None, None,
     None, None, None, None, datetime.date(2026, 8, 3)],
    ["测试原告乙", "测试被告B", "TEST-ACCOUNT-002", "test-pass-002", "立案成功", datetime.date(2026, 7, 22),
     "（2026）京0000民初00001号", "=_xlfn.DISPIMG(\"ID_TEST_001\",1)", None, None, None, datetime.date(2026, 8, 3)],
    ["测试原告丙", "测试被告C", "TEST-ACCOUNT-003", "test-pass-003", "已驳回", None, None,
     None, datetime.date(2026, 7, 28), "测试用驳回原因示例，请补充材料。", None, datetime.date(2026, 8, 3)],
]
QZ_ROWS = [
    ["测试原告丁", "测试被告D", "TEST-ACCOUNT-004", "test-pass-004", "审核中", None, None,
     None, None, None, None, datetime.date(2026, 8, 3)],
    ["测试原告戊", "测试被告E", "TEST-ACCOUNT-005", "test-pass-005", "强执成功", datetime.date(2026, 6, 3),
     "（2026）京0000执00001号", None, None, None, None, datetime.date(2026, 8, 3)],
    ["测试原告己", "测试被告F", "TEST-ACCOUNT-006", "test-pass-006", "已驳回", None, None,
     None, datetime.date(2026, 7, 24), "测试用驳回原因示例，请提供收款账户。", None, datetime.date(2026, 8, 3)],
]

DATE_FMT = "mm-dd-yy"


def write_header(ws, row, headers):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="FF92D050")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 27


def write_rows(ws, start_row, rows):
    for i, row in enumerate(rows):
        r = start_row + i
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            if col in (6, 9, 12):  # F/I/L 日期列
                cell.number_format = DATE_FMT
        ws.row_dimensions[r].height = 28


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    write_header(ws, 1, HEADER_LII)
    write_rows(ws, 2, LII_ROWS)
    qz_header_row = len(LII_ROWS) + 1 + 5  # 立案末行 + 5 = 4 + 5 = 9
    write_header(ws, qz_header_row, HEADER_QZ)
    write_rows(ws, qz_header_row + 1, QZ_ROWS)

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUT)
    print(f"OK: {OUT}  (强执表头行 {qz_header_row})")


if __name__ == "__main__":
    main()
