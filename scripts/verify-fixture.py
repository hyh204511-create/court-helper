#!/usr/bin/env python3
"""校验脱敏 fixture 结构（表头/样式/块行号/日期格式/DISPIMG）。

用法: python scripts/verify-fixture.py
失败时退出码非 0（供提交前检查）。
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

FIXTURE = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "立案与强执查询表-脱敏模板.xlsx"

EXPECTED = {
    "header_l1": ["原告", "被告", "账号", "密码", "立案状态", "立案成功时间", "案号",
                  "成功图片", "驳回时间", "驳回原因", "驳回图片", "查询时间"],
    "header_qz_row": 9,
    "header_qz": ["原告", "被告", "账号", "密码", "强执状态", "强执成功时间", "强执案号",
                  "成功图片", "驳回时间", "驳回原因", "驳回图片", "查询时间"],
    "bold": True,
    "fill": "FF92D050",
    "row_h_header": 27,
    "row_h_data": 28,
    "date_fmt": "mm-dd-yy",
    "dispimg_cell": "H3",
}


def main():
    errors = []
    wb = load_workbook(FIXTURE)
    if wb.sheetnames != ["Sheet1"]:
        errors.append(f"sheetnames={wb.sheetnames}")
    ws = wb["Sheet1"]

    def check(cond, msg):
        if not cond:
            errors.append(msg)

    for col, text in enumerate(EXPECTED["header_l1"], start=1):
        check(ws.cell(row=1, column=col).value == text, f"A{col}1 header={ws.cell(row=1, column=col).value!r}")
    for col, text in enumerate(EXPECTED["header_qz"], start=1):
        r = EXPECTED["header_qz_row"]
        check(ws.cell(row=r, column=col).value == text, f"row{r} header col{col}={ws.cell(row=r, column=col).value!r}")

    check(ws["A1"].font.bold is True, f"A1 bold={ws['A1'].font.bold}")
    fill = ws["A1"].fill.fgColor.rgb if ws["A1"].fill and ws["A1"].fill.patternType else None
    check(fill == EXPECTED["fill"], f"A1 fill={fill}")
    check(ws.row_dimensions[1].height == EXPECTED["row_h_header"], f"row1 height={ws.row_dimensions[1].height}")
    check(ws.row_dimensions[2].height == EXPECTED["row_h_data"], f"row2 height={ws.row_dimensions[2].height}")
    check(ws["F2"].number_format == EXPECTED["date_fmt"], f"F2 fmt={ws['F2'].number_format!r}")
    check(ws["L2"].number_format == EXPECTED["date_fmt"], f"L2 fmt={ws['L2'].number_format!r}")
    check(ws[EXPECTED["dispimg_cell"]].value and "DISPIMG" in str(ws[EXPECTED["dispimg_cell"]].value),
          f"{EXPECTED['dispimg_cell']} 无 DISPIMG: {ws[EXPECTED['dispimg_cell']].value!r}")
    check(ws["A2"].value == "测试原告甲", f"A2={ws['A2'].value!r}（应为模拟数据）")
    check(ws["C2"].value.startswith("TEST-"), f"C2={ws['C2'].value!r}（应为模拟账号）")

    if errors:
        print("FAIL:")
        for e in errors:
            print(" -", e)
        sys.exit(1)
    print("OK: fixture 结构校验通过")


if __name__ == "__main__":
    main()
