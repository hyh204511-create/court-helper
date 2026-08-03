#!/usr/bin/env python3
"""校验导出 xlsx 的图片锚点与单元格值（openpyxl 读回 + editAs 兼容）。

用法: python scripts/verify-export.py <xlsx> [--cells A1,E2,G3]
输出 JSON: {"sheets": [...], "media_count": N, "image_count": N,
            "anchors": [{"col": c, "row": r}, ...], "cells": {"A1": "原告", ...}}
无图片或图片与 media 文件数不一致时退出码非 0（带图验证失败）。

单元格值格式化: datetime → YYYY-MM-DD；None → ""；其余原样字符串。
--cells 未指定时不含 cells 字段。

兼容说明: openpyxl 3.1.5 解析带 editAs 属性的 oneCellAnchor/twoCellAnchor 会抛
TypeError（_AnchorBase.__init__ 不接受该参数），而 ExcelJS 生成的锚点默认带
editAs="oneCell"。本脚本在交给 openpyxl 解析前先剥离该属性；editAs 只影响
编辑器拖拽行为，不影响图片位置与内容。
"""
import argparse
import json
import re
import sys
import zipfile

from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.xml.functions import fromstring


def parse_drawing_anchors(path):
    """解析所有 drawing 中图片锚点（0 基 col/row）。"""
    with zipfile.ZipFile(path) as z:
        drawing_names = sorted(
            n for n in z.namelist() if re.match(r"xl/drawings/drawing\d+\.xml$", n)
        )
        anchors = []
        for dn in drawing_names:
            src = z.read(dn).decode("utf-8")
            src = re.sub(r'\seditAs="[^"]*"', "", src)  # openpyxl 3.1.5 兼容
            drawing = SpreadsheetDrawing.from_tree(fromstring(src.encode("utf-8")))
            for rel in drawing._blip_rels:
                frm = rel.anchor._from
                anchors.append({"col": frm.col, "row": frm.row})
        return anchors


def cell_value(v):
    if v is None:
        return ""
    if hasattr(v, "isoformat"):  # datetime/date
        return v.strftime("%Y-%m-%d")
    return str(v)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--cells", default="", help="逗号分隔单元格坐标，如 A1,E2,G3")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx)
    sheets = [ws.title for ws in wb.worksheets]
    cells = {}
    if args.cells:
        ws = wb.worksheets[0]
        for coord in args.cells.split(","):
            coord = coord.strip()
            if coord:
                cells[coord] = cell_value(ws[coord].value)
    with zipfile.ZipFile(args.xlsx) as z:
        media_count = len([n for n in z.namelist() if n.startswith("xl/media/") and not n.endswith("/")])
    anchors = parse_drawing_anchors(args.xlsx)
    out = {"sheets": sheets, "media_count": media_count, "image_count": len(anchors), "anchors": anchors}
    if args.cells:
        out["cells"] = cells
    print(json.dumps(out, ensure_ascii=False))
    ok = out["image_count"] > 0 and out["image_count"] == out["media_count"]
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
